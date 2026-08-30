"""Result rows and report writing."""

import csv
import datetime
import os
from dataclasses import asdict, dataclass, field
from typing import Optional

PASS = "PASS"
FAIL = "FAIL"
SKIP = "SKIP"
ERROR = "ERROR"

COLUMNS = [
    ("app_name", "App Name"),
    ("process_name", "Process"),
    ("round_number", "Round"),
    ("mode", "Launch Mode"),
    ("expected_mode", "Expected Mode"),
    ("detected_mode", "Detected Mode"),
    ("result", "Result"),
    ("switch_latency_s", "Switch Latency (s)"),
    ("deassert_latency_s", "De-assert Latency (s)"),
    ("workload_value", "Workload Hint"),
    ("power_source", "Power Source"),
    ("temperature", "Temp (C)"),
    ("pl1_max", "PL1MAX"),
    ("pl1_min", "PL1MIN"),
    ("reason", "Failure Reason"),
    ("notes", "Notes"),
    ("timestamp", "Timestamp"),
]


@dataclass
class ResultRow:
    app_name: str = ""
    process_name: str = ""
    round_number: int = 0
    mode: str = ""
    expected_mode: str = ""
    detected_mode: str = ""
    result: str = ""
    switch_latency_s: Optional[float] = None
    deassert_latency_s: Optional[float] = None
    workload_value: str = ""
    power_source: str = ""
    temperature: str = ""
    pl1_max: str = ""
    pl1_min: str = ""
    reason: str = ""
    notes: str = ""
    timestamp: str = field(
        default_factory=lambda: datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    def as_dict(self):
        return asdict(self)


def _sorted_for_report(rows):
    """Failures first, so the list of applications to fix is at the top."""
    priority = {FAIL: 0, ERROR: 1, SKIP: 2, PASS: 3}
    return sorted(
        rows,
        key=lambda row: (priority.get(row.result, 4), row.app_name.lower(),
                         row.round_number),
    )


def _format(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return "{0:.2f}".format(value)
    return value


def write_csv(rows, path):
    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([title for _key, title in COLUMNS])
        for row in _sorted_for_report(rows):
            data = row.as_dict()
            writer.writerow([_format(data[key]) for key, _title in COLUMNS])
    return path


def write_xlsx(rows, path):
    """Write an Excel report with an auto-filter and highlighted failures.

    Returns None when openpyxl is unavailable, which is normal for a build that
    was packaged without it; the CSV is always written regardless.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0068B5")
    fills = {
        FAIL: PatternFill("solid", fgColor="FFD5D5"),
        ERROR: PatternFill("solid", fgColor="FFE6CC"),
        SKIP: PatternFill("solid", fgColor="EFEFEF"),
    }

    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    results = workbook.active
    results.title = "Results"
    results.append(["#", "application", "APAT results", "pass/fail"])
    for cell in results[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for entry in simple_rows(rows):
        results.append([entry["number"], entry["application"],
                        entry["apat_result"], entry["verdict"]])
        row_fill = {"pass": pass_fill, "fail": fail_fill}.get(entry["verdict"])
        if row_fill is not None:
            for cell in results[results.max_row]:
                cell.fill = row_fill
    results.freeze_panes = "A2"
    for index, width in enumerate((6, 30, 26, 12), start=1):
        results.column_dimensions[get_column_letter(index)].width = width

    sheet = workbook.create_sheet("Details")
    sheet.append([title for _key, title in COLUMNS])
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in _sorted_for_report(rows):
        data = row.as_dict()
        sheet.append([_format(data[key]) for key, _title in COLUMNS])
        fill = fills.get(row.result)
        if fill is not None:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:{0}{1}".format(
        get_column_letter(len(COLUMNS)), max(sheet.max_row, 1)
    )
    for index, (key, title) in enumerate(COLUMNS, start=1):
        longest = max(
            [len(title)] + [len(str(_format(row.as_dict()[key]))) for row in rows] or [0]
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 60)

    summary = workbook.create_sheet("Summary")
    summary.append(["App Name", "Process", "Expected Mode", "Rounds", "Pass", "Fail",
                    "Skip/Error", "Verdict"])
    for cell in summary[1]:
        cell.font = header_font
        cell.fill = header_fill

    for line in summarize(rows):
        summary.append([
            line["app_name"], line["process_name"], line["expected_mode"],
            line["rounds"], line["passed"], line["failed"], line["other"],
            line["verdict"],
        ])
    for index in range(1, 9):
        summary.column_dimensions[get_column_letter(index)].width = 18

    workbook.save(path)
    return path


def summarize(rows):
    """Per-application roll-up; flags apps that only fail some of the time."""
    order = []
    grouped = {}
    for row in rows:
        key = row.process_name or row.app_name
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(row)

    summary = []
    for key in order:
        entries = grouped[key]
        passed = sum(1 for r in entries if r.result == PASS)
        failed = sum(1 for r in entries if r.result == FAIL)
        other = sum(1 for r in entries if r.result in (SKIP, ERROR))
        if failed == 0 and passed:
            verdict = "PASS"
        elif passed == 0 and failed:
            verdict = "FAIL"
        elif failed and passed:
            verdict = "INTERMITTENT"
        else:
            verdict = "NOT TESTED"
        summary.append({
            "app_name": entries[0].app_name,
            "process_name": entries[0].process_name,
            "expected_mode": entries[0].expected_mode,
            "rounds": len(entries),
            "passed": passed,
            "failed": failed,
            "other": other,
            "verdict": verdict,
        })
    return summary


def simple_rows(rows):
    """The compact per-application view: one row, one verdict.

    Rounds collapse into a single line: the reported action set is the one from
    a failing round when there is one, so the table shows what actually went
    wrong rather than a later round that happened to pass.
    """
    order = []
    grouped = {}
    for row in rows:
        key = row.process_name or row.app_name
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(row)

    output = []
    for number, key in enumerate(order, start=1):
        entries = grouped[key]
        failures = [r for r in entries if r.result == FAIL]
        skipped = [r for r in entries if r.result in (SKIP, ERROR)]

        if failures:
            representative, verdict = failures[0], "fail"
        elif len(skipped) == len(entries):
            representative, verdict = entries[0], "skip"
        else:
            representative = next(r for r in entries if r.result == PASS)
            verdict = "pass"

        output.append({
            "number": number,
            "application": representative.process_name or representative.app_name,
            "apat_result": representative.detected_mode or representative.reason or "-",
            "verdict": verdict,
        })
    return output


def write_simple_csv(rows, path):
    """The report layout asked for: number, application, APAT result, pass/fail."""
    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["#", "application", "APAT results", "pass/fail"])
        for entry in simple_rows(rows):
            writer.writerow([entry["number"], entry["application"],
                             entry["apat_result"], entry["verdict"]])
    return path


def timestamped_paths(output_dir, formats):
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    paths = {}
    for fmt in formats:
        paths[fmt] = os.path.join(output_dir, "dtt_wl_report_{0}.{1}".format(stamp, fmt))
    return paths
